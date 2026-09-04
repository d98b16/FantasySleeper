import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---------- palette ----------
NAVY = "1F3864"
WHITE = "FFFFFF"
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(size=10, bold=False, italic=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, italic=italic, color=color)

def fill(hexcol):
    return PatternFill("solid", fgColor=hexcol)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFTTOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

TIER_FILL = {
    1: "FFD966", 2: "A9D08E", 3: "9DC3E6", 4: "B4A7D6", 5: "F4B183",
    6: "C9DAF8", 7: "D5A6BD", 8: "A2C4C9", 9: "F6B26B", 10: "D9D9D9",
}
POS_COLOR = {"RB": "137333", "WR": "1155CC", "QB": "B45F06", "TE": "6A329F",
             "DST": "666666", "K": "666666"}

# ============================================================
# TAB 1 — DRAFT PLAN (YOU @ 3)
# ============================================================
ws = wb.active
ws.title = "Draft Plan (You @ 3)"
ws.sheet_properties.tabColor = NAVY

ws["A1"] = "2026 FANTASY DRAFT CHEAT SHEET  —  Pick #3, Snake"
ws["A1"].font = F(16, bold=True, color=NAVY)
ws.merge_cells("A1:F1")

banner = ("VERIFIED FROM SLEEPER — \"GovSmart Gridiron\":  12-team snake  •  0.5 PPR  •  6-PT passing TDs  •  $100 FAAB  •  your slot #3  •  "
          "draft Fri Sep 4, 5:30 PM ET, 90s pick timer.   "
          "ROSTER: QB / RB / RB / WR / WR / TE / FLEX / DEF + 4 BENCH — only 12 rounds, NO KICKER, and a DEFENSE is a required starter. "
          "2 IR slots let you stash an injured guy. "
          "6-pt pass TDs lift EVERY QB including the streamer, so the elite QB's edge grows only 7-19 pts ALL SEASON "
          "(measured across 2024+2025) — WAIT ON QB until round 9+, then take the highest PASSING-TD volume left "
          "(the rule pays 2 pts per passing TD and nothing for rushing).")
ws["A2"] = banner
ws["A2"].font = F(9, italic=True, color="B45309")
ws["A2"].fill = fill("FFF2CC")
ws["A2"].alignment = LEFT
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 56

ws["A3"] = ("YOUR 12 PICKS (slot 3, snake):  R1 #3  •  R2 #22  •  R3 #27  •  R4 #46  •  R5 #51  •  R6 #70  •  "
            "R7 #75  •  R8 #94  •  R9 #99  •  R10 #118  •  R11 #123  •  R12 #142   —  that's the whole draft.")
ws["A3"].font = F(9, bold=True, color="333333")
ws["A3"].alignment = LEFT
ws.merge_cells("A3:F3")
ws.row_dimensions[3].height = 28

headers = ["Rd", "Your Pick", "On Board (ADP)", "BEST TARGET (take one)", "Fallbacks / Best Available", "Plan"]
hrow = 5
for c, h in enumerate(headers, 1):
    cell = ws.cell(hrow, c, h)
    cell.font = F(11, bold=True, color=WHITE)
    cell.fill = fill(NAVY)
    cell.alignment = CENTER
    cell.border = BORDER

plan = [
    ("1", "3", "1-6", "Ja'Marr Chase or Puka Nacua (elite WR1)", "Jaxon Smith-Njigba; Bijan Robinson if he slips", "Gibbs & Bijan usually gone 1-2. Lock an elite WR. Don't overthink it."),
    ("2", "22", "18-25", "Kenneth Walker III or Omarion Hampton (anchor an RB)", "Ashton Jeanty if he slides (ankle discount); Nico Collins / A.J. Brown (WR)", "RB dries up fast. Grab your RB1 at the turn."),
    ("3", "27", "24-31", "Malik Nabers or Zay Flowers (WR); Kyren Williams (RB)", "Chris Olave; Jeremiyah Love; DeVonta Smith", "You pick 22 & 27 back-to-back. Take whichever position you skipped at 22."),
    ("4", "46", "43-52", "Brock Bowers (elite-TE edge) OR Terry McLaurin / DJ Moore (WR)", "Bucky Irving; Quinshon Judkins (RB)", "Only 12 rounds — Bowers here locks TE and frees every later pick."),
    ("5", "51", "48-58", "RB Quinshon Judkins or WR Mike Evans — keep stacking starters", "Bucky Irving; Rome Odunze; Jaylen Waddle", "Do NOT take a QB here. Top QB is worth ~60 pts over replacement; a top RB is worth ~227."),
    ("6", "70", "60-72", "DK Metcalf / Marvin Harrison Jr. (WR upside)", "TreVeyon Henderson (RB); Christian Watson", "Fill your last open starting WR or RB here. Still not QB."),
    ("7", "75", "72-80", "Jordan Addison / Jayden Reed (WR)", "Tony Pollard (RB); a starting TE if you still need one", "Bench is only 4 deep — starters beat lottery tickets."),
    ("8", "94", "87-99", "Finish your 7 offensive starters — best WR/RB/TE left", "Tyler Warren / Kyle Pitts (TE); Matthew Golden (WR)", "Every starter filled before you spend a pick on QB."),
    ("9", "99", "94-104", "QB NOW — Stafford / Prescott / Herbert / Lawrence / Mahomes", "Best FLEX left if two QBs you like are still there", "The QB window. 2025: Stafford +92 pts from the 6-pt rule vs Allen +50, at 56 picks cheaper."),
    ("10", "118", "110-123", "Handcuff your RB1/RB2 (RJ Harvey, Jonathon Brooks)", "Kyle Pitts (TE2); Justin Herbert / Bo Nix (QB2)", "Only 4 bench spots — protect the studs you already paid for."),
    ("11", "123", "Best avail.", "Best upside swing left on the board", "2nd TE or QB if you're thin there", "Last real value pick. Ceiling over floor."),
    ("12", "142", "Last pick", "DEFENSE — Seattle / Denver / Houston (best Week-1 matchup)", "—", "NO KICKER in this league. DEF is a required starter — this is the pick for it."),
]
r = hrow + 1
for row in plan:
    for c, val in enumerate(row, 1):
        cell = ws.cell(r, c, val)
        cell.border = BORDER
        if c == 1:
            cell.font = F(11, bold=True, color=NAVY); cell.alignment = CENTER
        elif c == 2:
            cell.font = F(11, bold=True, color="000000"); cell.alignment = CENTER
        elif c == 3:
            cell.font = F(9, color="555555"); cell.alignment = CENTER
        elif c == 4:
            cell.font = F(10, bold=True); cell.alignment = LEFT; cell.fill = fill("EAF3EA")
        else:
            cell.font = F(9); cell.alignment = LEFT
    ws.row_dimensions[r].height = 40
    r += 1

# ---- turn scenarios (decision tree) ----
scenarios = [
    "DEFAULT (you took a WR at 3): leave the 22/27 turn with an RB anchor + one more stud.  Ideal pairs: Walker + Nabers  ·  Hampton + Kyren  ·  Jacobs + Olave.",
    "CHASE AND NACUA BOTH GONE AT 3: take Jaxon Smith-Njigba. Do NOT reach for an RB at 3 — the Tier-1 WR still wins the pick.",
    "GIBBS OR BIJAN SLIDES TO 3: take him. Then go WR + WR at 22/27 (A.J. Brown / Nico + Nabers / Flowers) — your RB1 is already banked.",
    "JEANTY FALLS TO 22: ankle discount, not a red flag — and you have 2 IR slots, so a slow start is survivable. Otherwise Walker is the safe anchor.",
    "ROSTER SHAPE: 12 rounds, 8 starters (QB/RB/RB/WR/WR/TE/FLEX/DEF), 4 bench, NO KICKER. Every pick is a roster spot — there is no room for dead weight.",
    "QB, MEASURED (nflverse 2024+2025, re-scored to 6-pt): the rule lifts the STREAMER from 17.4 to 20.9 pts/g and the top-6 from 20.7 to 24.5 — the elite QB's ADVANTAGE grows only 7-19 pts across a WHOLE SEASON (median 12.5 over 2 seasons x 3 replacement ranks). For scale, value over replacement: best QB ~60, best TE ~113, best WR ~117, best RB ~227. The rule pays exactly 2 pts per PASSING TD and nothing for rushing, so judge a QB on passing-TD volume, not on being a dual-threat. Wait until round 9.",
]
sr = r + 1
sb = ws.cell(sr, 1, "TURN SCENARIOS — PICKS 22 & 27 (you pick back-to-back)")
sb.font = F(11, bold=True, color=WHITE); sb.fill = fill("CC0000"); sb.alignment = LEFT
ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=6)
for s in scenarios:
    sr += 1
    c = ws.cell(sr, 1, s); c.font = F(9); c.alignment = LEFT; c.fill = fill("FBEFEF")
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=6)
    ws.row_dimensions[sr].height = 26

widths = {"A": 4, "B": 8, "C": 12, "D": 40, "E": 34, "F": 40}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A6"

# ============================================================
# TAB 2 — BIG BOARD (Top ~120)
# ============================================================
wb2 = wb.create_sheet("Big Board")
wb2.sheet_properties.tabColor = "1155CC"

# board rows: (player, pos, team, bye, adp, tier, notes)   adp="" when beyond live-ADP range
# ADP refreshed 2026-08-27 (FantasyFootballCalculator half-PPR)
board = [
    ("Jahmyr Gibbs","RB","DET",6,"1.5",1,""),
    ("Bijan Robinson","RB","ATL",11,"2.2",1,""),
    ("Puka Nacua","WR","LAR",11,"3.0",1,"Groin — minor, expected Wk1"),
    ("Ja'Marr Chase","WR","CIN",6,"3.9",1,"Safest WR1"),
    ("Jaxon Smith-Njigba","WR","SEA",11,"5.2",1,""),
    ("Jonathan Taylor","RB","IND",13,"5.9",2,""),
    ("Christian McCaffrey","RB","SF",8,"6.6",2,"Elite when healthy"),
    ("Amon-Ra St. Brown","WR","DET",6,"8.0",2,""),
    ("James Cook III","RB","BUF",7,"9.3",2,""),
    ("Derrick Henry","RB","BAL",13,"9.7",2,"Age, still elite"),
    ("De'Von Achane","RB","MIA",6,"11.7",2,""),
    ("Drake London","WR","ATL",11,"12.0",2,""),
    ("CeeDee Lamb","WR","DAL",14,"12.5",2,""),
    ("Justin Jefferson","WR","MIN",6,"14.4",3,""),
    ("Chase Brown","RB","CIN",6,"15.1",3,""),
    ("Rashee Rice","WR","KC",5,"15.2",3,""),
    ("Saquon Barkley","RB","PHI",10,"17.6",3,""),
    ("Nico Collins","WR","HOU",8,"18.4",3,""),
    ("Ashton Jeanty","RB","LV",13,"18.9",3,"Ankle — Wk1 uncertain; ADP sliding (~15 to 19)"),
    ("George Pickens","WR","DAL",14,"19.9",3,""),
    ("A.J. Brown","WR","NE",11,"20.0",3,"New team (NE)"),
    ("Kenneth Walker III","RB","KC",5,"20.5",3,"Now in KC"),
    ("Chris Olave","WR","NO",8,"23.2",4,""),
    ("Omarion Hampton","RB","LAC",7,"24.0",4,""),
    ("Zay Flowers","WR","BAL",13,"25.2",4,""),
    ("Josh Jacobs","RB","GB",11,"25.4",4,""),
    ("Malik Nabers","WR","NYG",8,"26.2",4,"Knee — expected Wk1"),
    ("Kyren Williams","RB","LAR",11,"27.4",4,""),
    ("Jeremiyah Love","RB","ARI",14,"29.0",4,"Rookie; high-ankle — expected Wk1"),
    ("Tetairoa McMillan","WR","CAR",5,"30.7",4,""),
    ("Garrett Wilson","WR","NYJ",13,"31.1",4,""),
    ("DeVonta Smith","WR","PHI",10,"31.6",4,"Value"),
    ("Josh Allen","QB","BUF",7,"33.0",5,"QB1 overall — 6-pt pass-TD riser"),
    ("Javonte Williams","RB","DAL",14,"33.2",5,"Now in DAL"),
    ("Breece Hall","RB","NYJ",13,"33.5",5,"Groin — expected Wk1; market rising"),
    ("Emeka Egbuka","WR","TB",10,"33.7",5,"Toe — monitor (Q)"),
    ("Tee Higgins","WR","CIN",6,"35.9",5,""),
    ("Travis Etienne Jr.","RB","NO",8,"37.3",5,"Now in NO"),
    ("Trey McBride","TE","ARI",14,"37.8",5,"TE1 — elite"),
    ("Davante Adams","WR","LAR",11,"37.9",5,"Now in LAR"),
    ("Cam Skattebo","RB","NYG",8,"38.4",5,"Risky at cost"),
    ("Jameson Williams","WR","DET",6,"38.7",5,"Upside"),
    ("Ladd McConkey","WR","LAC",7,"41.0",5,""),
    ("D'Andre Swift","RB","CHI",10,"43.2",6,""),
    ("Brock Bowers","TE","LV",13,"44.0",6,"Elite TE — positional edge"),
    ("DJ Moore","WR","BUF",7,"44.9",6,"Now in BUF"),
    ("Terry McLaurin","WR","WAS",7,"45.1",6,""),
    ("Bucky Irving","RB","TB",10,"47.0",6,""),
    ("Jaylen Waddle","WR","DEN",10,"47.3",6,"Now in DEN"),
    ("Rome Odunze","WR","CHI",10,"48.9",6,""),
    ("Drake Maye","QB","NE",11,"51.3",6,"6-pt pass-TD riser"),
    ("Quinshon Judkins","RB","CLE",11,"51.5",6,""),
    ("Bhayshul Tuten","RB","JAX",7,"52.5",6,"Upside"),
    ("Mike Evans","WR","SF",8,"54.2",6,"Now in SF"),
    ("Lamar Jackson","QB","BAL",13,"55.5",6,"6-pt pass-TD riser (rush+pass)"),
    ("David Montgomery","RB","HOU",8,"55.5",6,"Now in HOU"),
    ("Christian Watson","WR","GB",11,"56.0",7,""),
    ("Joe Burrow","QB","CIN",6,"57.5",7,"6-pt pass-TD riser"),
    ("Luther Burden III","WR","CHI",10,"57.6",7,"Groin — expected Wk1; ADP rising"),
    ("Courtland Sutton","WR","DEN",10,"59.1",7,""),
    ("Ollie Gordon II","RB","MIA",6,"",7,"Fell out of top-60 ADP — monitor role"),
    ("Alec Pierce","WR","IND",13,"",7,"Ankle/PUP — expected Wk1"),
    ("DK Metcalf","WR","PIT",9,"",7,""),
    ("Marvin Harrison Jr.","WR","ARI",14,"",7,"Bounce-back upside"),
    ("Parker Washington","WR","JAX",7,"",7,""),
    ("Brian Thomas Jr.","WR","JAX",7,"",7,"Upside"),
    ("Jaylen Warren","RB","PIT",9,"",7,""),
    ("Carnell Tate","WR","TEN",9,"",7,"Rookie"),
    ("Rhamondre Stevenson","RB","NE",11,"",7,""),
    ("Chris Godwin Jr.","WR","TB",10,"",7,""),
    ("TreVeyon Henderson","RB","NE",11,"",7,"Upside"),
    ("Michael Pittman Jr.","WR","PIT",9,"",8,"Now in PIT"),
    ("Tony Pollard","RB","TEN",9,"",8,""),
    ("Jakobi Meyers","WR","JAX",7,"",8,"Now in JAX"),
    ("Jordan Addison","WR","MIN",6,"",8,""),
    ("Quentin Johnston","WR","LAC",7,"",8,""),
    ("Jayden Reed","WR","GB",11,"",8,""),
    ("George Kittle","TE","SF",8,"",8,"Achilles — expected Wk1"),
    ("Wan'Dale Robinson","WR","TEN",9,"",8,"Now in TEN"),
    ("Josh Downs","WR","IND",13,"",8,""),
    ("Rico Dowdle","RB","PIT",9,"",8,""),
    ("Dak Prescott","QB","DAL",14,"",8,""),
    ("Stefon Diggs","WR","WAS",7,"",8,"Now in WAS"),
    ("Xavier Worthy","WR","KC",5,"",8,""),
    ("Deebo Samuel Sr.","WR","SF",8,"",8,""),
    ("J.K. Dobbins","RB","DEN",10,"",8,""),
    ("Jayden Daniels","QB","WAS",7,"",8,"6-pt pass-TD riser (rush+pass)"),
    ("Jadarian Price","RB","SEA",11,"",8,"Rookie"),
    ("Matthew Stafford","QB","LAR",11,"",8,""),
    ("Matthew Golden","WR","GB",11,"",9,"Upside"),
    ("Colston Loveland","TE","CHI",10,"",9,""),
    ("Jalen Hurts","QB","PHI",10,"",9,"Rushing QB — 6-pt pass-TD riser"),
    ("Brock Purdy","QB","SF",8,"",9,""),
    ("Chuba Hubbard","RB","CAR",5,"",9,"Hamstring; committee w/ Brooks"),
    ("Khalil Shakir","WR","BUF",7,"",9,""),
    ("Romeo Doubs","WR","NE",11,"",9,""),
    ("Makai Lemon","WR","PHI",10,"",9,"Rookie — upside"),
    ("Trevor Lawrence","QB","JAX",7,"",9,""),
    ("Caleb Williams","QB","CHI",10,"",9,"Upside"),
    ("Jared Goff","QB","DET",6,"",9,""),
    ("Tyler Warren","TE","IND",13,"",9,"Groin — expected Wk1"),
    ("Patrick Mahomes","QB","KC",5,"",9,"Knee — reduced rush upside"),
    ("Jalen Coker","WR","CAR",5,"",9,""),
    ("Rashid Shaheed","WR","SEA",11,"",9,"Now in SEA"),
    ("Jerry Jeudy","WR","CLE",11,"",9,""),
    ("KC Concepcion","WR","CLE",11,"",9,"Rookie"),
    ("Harold Fannin Jr.","TE","CLE",11,"",9,""),
    ("Jonathon Brooks","RB","CAR",5,"",9,"Committee w/ Hubbard"),
    ("Justin Herbert","QB","LAC",7,"",10,""),
    ("Kyle Pitts Sr.","TE","ATL",11,"",10,""),
    ("Bo Nix","QB","DEN",10,"",10,""),
    ("Jaxson Dart","QB","NYG",8,"",10,"Upside"),
    ("Kyle Monangai","RB","CHI",10,"",10,"Knee — multi-week (Q)"),
    ("Dallas Goedert","TE","PHI",10,"",10,""),
    ("Kenny Gainwell","RB","TB",10,"",10,""),
    ("Tre Tucker","WR","LV",13,"",10,""),
    ("RJ Harvey","RB","DEN",10,"",10,"Handcuff (Dobbins)"),
    ("HOU D/ST","DST","HOU",8,"",10,"DEF #1 in 2025 (146 pts, 8.6/g) — REQUIRED starter, take one by R11"),
    ("SEA D/ST","DST","SEA",11,"",10,"DEF #2 in 2025 (140 pts, 8.2/g) — REQUIRED starter, take one by R11"),
    ("DEN D/ST","DST","DEN",10,"",10,"DEF #3 in 2025 (137 pts, 8.1/g) — REQUIRED starter, take one by R11"),
    ("MIN D/ST","DST","MIN",6,"",10,"DEF #4 in 2025 (129 pts, 7.6/g) — REQUIRED starter, take one by R11"),
    ("LAR D/ST","DST","LAR",11,"",10,"DEF #5 in 2025 (128 pts, 7.5/g) — REQUIRED starter, take one by R11"),
    ("JAX D/ST","DST","JAX",7,"",10,"DEF #6 in 2025 (128 pts, 7.5/g) — REQUIRED starter, take one by R11"),
    ("CLE D/ST","DST","CLE",11,"",10,"DEF #7 in 2025 (127 pts, 7.5/g) — REQUIRED starter, take one by R11"),
    ("PHI D/ST","DST","PHI",10,"",10,"DEF #8 in 2025 (125 pts, 7.3/g) — REQUIRED starter, take one by R11"),
    ("ATL D/ST","DST","ATL",11,"",10,"DEF #9 in 2025 (117 pts, 6.9/g) — REQUIRED starter, take one by R11"),
    ("PIT D/ST","DST","PIT",9,"",10,"DEF #10 in 2025 (117 pts, 6.9/g) — REQUIRED starter, take one by R11"),
    ("CHI D/ST","DST","CHI",10,"",10,"DEF #11 in 2025 (114 pts, 6.7/g) — REQUIRED starter, take one by R11"),
    ("LAC D/ST","DST","LAC",7,"",10,"DEF #12 in 2025 (108 pts, 6.3/g) — REQUIRED starter, take one by R11"),
]

wb2["A1"] = "BIG BOARD — 2026 Half-PPR (Top 120)"
wb2["A1"].font = F(16, bold=True, color=NAVY)
wb2.merge_cells("A1:I1")
wb2["A2"] = ("Drafted? (col I): x for EVERY pick by anyone — greys the row out.  MINE? (col J): x for YOUR picks — turns green and feeds the Live Tracker tab.  "
             "ADP = live 12-team half-PPR market thru ~#60 (refreshed Aug 27); 61+ = expert consensus.  QBs carry a 6-pt pass-TD bump.")
wb2["A2"].font = F(9, italic=True, color="555555")
wb2["A2"].alignment = LEFT
wb2.merge_cells("A2:I2")
wb2.row_dimensions[2].height = 30

bh = ["Rank","Tier","Player","Pos","Team","Bye","ADP","Notes","Drafted?","MINE?"]
hrow2 = 4
for c, h in enumerate(bh, 1):
    cell = wb2.cell(hrow2, c, h)
    cell.font = F(11, bold=True, color=WHITE)
    cell.fill = fill(NAVY); cell.alignment = CENTER; cell.border = BORDER

r = hrow2 + 1
for i, (player, pos, team, bye, adp, tier, notes) in enumerate(board, 1):
    wb2.cell(r, 1, i).font = F(10, bold=True); wb2.cell(r, 1).alignment = CENTER
    tcell = wb2.cell(r, 2, "T%d" % tier); tcell.alignment = CENTER
    tcell.fill = fill(TIER_FILL[tier]); tcell.font = F(9, bold=True)
    wb2.cell(r, 3, player).font = F(10, bold=True)
    pcell = wb2.cell(r, 4, pos); pcell.alignment = CENTER
    pcell.font = F(10, bold=True, color=POS_COLOR.get(pos, "000000"))
    wb2.cell(r, 5, team).alignment = CENTER; wb2.cell(r, 5).font = F(10)
    wb2.cell(r, 6, bye).alignment = CENTER; wb2.cell(r, 6).font = F(10, color="777777")
    wb2.cell(r, 7, adp).alignment = CENTER; wb2.cell(r, 7).font = F(10)
    wb2.cell(r, 8, notes).alignment = LEFT; wb2.cell(r, 8).font = F(9, color="B45309" if notes else "000000")
    wb2.cell(r, 9, "").alignment = CENTER
    wb2.cell(r, 10, "").alignment = CENTER
    for c in range(1, 11):
        wb2.cell(r, c).border = BORDER
    r += 1
last = r - 1

bwidths = {"A":6,"B":6,"C":24,"D":6,"E":7,"F":6,"G":8,"H":34,"I":9,"J":8}
for col, w in bwidths.items():
    wb2.column_dimensions[col].width = w
wb2.freeze_panes = "A5"

# drafted dropdown + grey-out rule
dv = DataValidation(type="list", formula1='"x"', allow_blank=True)
wb2.add_data_validation(dv)
dv.add("I5:I%d" % last)
dv.add("J5:J%d" % last)
mine_green = fill("D9EAD3")
wb2.conditional_formatting.add(
    "A5:J%d" % last,
    FormulaRule(formula=['$J5="x"'], fill=mine_green, font=F(10, bold=True), stopIfTrue=True),
)
grey = fill("E8E8E8")
strike = F(10, italic=True, color="999999")
wb2.conditional_formatting.add(
    "A5:J%d" % last,
    FormulaRule(formula=['$I5="x"'], fill=grey, font=strike, stopIfTrue=False),
)

# ============================================================
# TAB 2.5 — LIVE TRACKER (formulas)
# ============================================================
from openpyxl.formatting.rule import CellIsRule
wbT = wb.create_sheet("Live Tracker")
wbT.sheet_properties.tabColor = "CC0000"
# Derived from the Big Board rows we actually wrote (hrow2+1 .. last), never
# hardcoded. A literal $5:$124 happens to be right at exactly 120 players and
# silently points at the wrong cells the moment the board changes length --
# the same class of bug that already bit the CSV once.
BB_FIRST, BB_LAST = hrow2 + 1, last
def _bb(col):
    return "'Big Board'!$%s$%d:$%s$%d" % (col, BB_FIRST, col, BB_LAST)
BB_POS, BB_TIER, BB_BYE, BB_DR, BB_MINE = (_bb(c) for c in "DBFIJ")
assert BB_LAST - BB_FIRST + 1 == len(board), \
    "Big Board range %d-%d covers %d rows but board has %d" % (
        BB_FIRST, BB_LAST, BB_LAST - BB_FIRST + 1, len(board))

wbT["A1"] = "LIVE DRAFT TRACKER — computes as you mark the Big Board"
wbT["A1"].font = F(15, bold=True, color=NAVY)
wbT.merge_cells("A1:J1")
wbT["A2"] = ("During the draft: on the Big Board put x in Drafted? (col I) for EVERY pick made by anyone, "
             "and ALSO x in MINE? (col J) for your own picks. 12 rounds, no kicker, DEF required in R12.")
wbT["A2"].font = F(9, italic=True, color="555555")
wbT["A2"].alignment = LEFT
wbT.merge_cells("A2:J2")
wbT.row_dimensions[2].height = 28

POSES = ["QB","RB","WR","TE"]
for j, p in enumerate(POSES):
    c = wbT.cell(4, 2+j, p); c.font = F(10, bold=True, color=WHITE); c.fill = fill(NAVY); c.alignment = CENTER; c.border = BORDER
wbT.cell(5,1,"My roster count").font = F(10, bold=True)
wbT.cell(6,1,"Starters still needed").font = F(10, bold=True)
wbT.cell(7,1,"Top-tier left (undrafted)").font = F(10, bold=True)
wbT.cell(8,1,"Next tier left").font = F(10, bold=True)
need = {"QB":1,"RB":2,"WR":2,"TE":1}
for j, p in enumerate(POSES):
    col = 2+j
    L = get_column_letter(col)
    wbT.cell(5,col,'=COUNTIFS(%s,%s$4,%s,"x")' % (BB_POS, L, BB_MINE))
    wbT.cell(6,col,'=MAX(0,%d-%s5)' % (need[p], L))
    _ts = sorted({b[5] for b in board if b[1] == p})
    _el, _md = _ts[:2], _ts[2:4]
    wbT.cell(7,col,"=" + "+".join('COUNTIFS(%s,%s$4,%s,"T%d",%s,"<>x")' % (BB_POS,L,BB_TIER,t,BB_DR) for t in _el))
    wbT.cell(8,col,"=" + "+".join('COUNTIFS(%s,%s$4,%s,"T%d",%s,"<>x")' % (BB_POS,L,BB_TIER,t,BB_DR) for t in _md))
for rr in range(5,9):
    for cc in range(1,6):
        wbT.cell(rr,cc).border = BORDER
        if cc >= 2:
            wbT.cell(rr,cc).alignment = CENTER
wbT.cell(6,6,"(+1 FLEX, +1 DEF in R12 — no kicker)").font = F(9, italic=True, color="777777")

wbT.cell(10,1,"My bye weeks:").font = F(10, bold=True)
BYES = [5,6,7,8,9,10,11,13,14]
for j,b in enumerate(BYES):
    c = wbT.cell(10,2+j,b); c.font = F(10, bold=True, color=WHITE); c.fill = fill("595959"); c.alignment = CENTER; c.border = BORDER
wbT.cell(11,1,"Count (3+ same bye = fix it)").font = F(10, bold=True)
for j,b in enumerate(BYES):
    L = get_column_letter(2+j)
    c = wbT.cell(11,2+j,'=COUNTIFS(%s,%s$10,%s,"x")' % (BB_BYE, L, BB_MINE))
    c.alignment = CENTER; c.border = BORDER
wbT.conditional_formatting.add("B11:J11", CellIsRule(operator="greaterThanOrEqual", formula=["3"], fill=fill("F4CCCC"), font=F(10, bold=True, color="CC0000")))

wbT.cell(13,1,"SNAKE PICK CALCULATOR").font = F(11, bold=True, color=NAVY)
wbT.cell(14,1,"Teams:").font = F(10)
tc = wbT.cell(14,2,12); tc.fill = fill("FFF2CC"); tc.font = F(10, bold=True); tc.alignment = CENTER; tc.border = BORDER
wbT.cell(15,1,"Your slot:").font = F(10)
sc = wbT.cell(15,2,3); sc.fill = fill("FFF2CC"); sc.font = F(10, bold=True); sc.alignment = CENTER; sc.border = BORDER
wbT.cell(16,1,"(yellow cells are inputs — picks below recompute if league changes)").font = F(8, italic=True, color="777777")
wbT.merge_cells("A16:E16")
for jj, h in enumerate(("Round","Overall pick")):
    hc = wbT.cell(17,1+jj,h)
    hc.font = F(10, bold=True, color=WHITE); hc.fill = fill(NAVY); hc.alignment = CENTER; hc.border = BORDER
for rd in range(1,13):
    rr = 17+rd
    a = wbT.cell(rr,1,rd); a.alignment = CENTER; a.border = BORDER
    b = wbT.cell(rr,2,'=IF(MOD(A%d,2)=1,$B$14*(A%d-1)+$B$15,$B$14*A%d-$B$15+1)' % (rr,rr,rr))
    b.alignment = CENTER; b.font = F(10, bold=True); b.border = BORDER

wbT.column_dimensions["A"].width = 27
for cl in "BCDEFGHIJ":
    wbT.column_dimensions[cl].width = 8
wbT.freeze_panes = "A3"

# ============================================================
# TAB 3 — BY POSITION
# ============================================================
wb3 = wb.create_sheet("By Position")
wb3.sheet_properties.tabColor = "137333"

def pos_block(ws, start_col, title, color, players):
    # players: list of (tier, name, team, adp)
    c0 = start_col
    tcell = ws.cell(1, c0, title)
    tcell.font = F(12, bold=True, color=WHITE)
    tcell.fill = fill(color); tcell.alignment = CENTER
    ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0+3)
    sub = ["Tier","Player","Tm","ADP"]
    for j, s in enumerate(sub):
        cell = ws.cell(2, c0+j, s); cell.font = F(9, bold=True, color=WHITE)
        cell.fill = fill("595959"); cell.alignment = CENTER; cell.border = BORDER
    rr = 3
    for (tier, name, team, adp) in players:
        tc = ws.cell(rr, c0, "T%d" % tier); tc.alignment = CENTER
        tc.fill = fill(TIER_FILL[tier]); tc.font = F(8, bold=True); tc.border = BORDER
        nc = ws.cell(rr, c0+1, name); nc.font = F(9, bold=True); nc.border = BORDER
        tmc = ws.cell(rr, c0+2, team); tmc.alignment = CENTER; tmc.font = F(9); tmc.border = BORDER
        ac = ws.cell(rr, c0+3, adp); ac.alignment = CENTER; ac.font = F(9, color="555555"); ac.border = BORDER
        rr += 1
    # widths
    ws.column_dimensions[get_column_letter(c0)].width = 5
    ws.column_dimensions[get_column_letter(c0+1)].width = 21
    ws.column_dimensions[get_column_letter(c0+2)].width = 5
    ws.column_dimensions[get_column_letter(c0+3)].width = 7

wb3["A1"]  # ensure exists
RB = [(1,"Jahmyr Gibbs","DET","1.5"),(1,"Bijan Robinson","ATL","2.2"),(2,"Jonathan Taylor","IND","5.9"),
      (2,"C. McCaffrey","SF","6.6"),(2,"James Cook III","BUF","9.3"),(2,"Derrick Henry","BAL","9.7"),
      (2,"De'Von Achane","MIA","11.7"),(3,"Chase Brown","CIN","15.1"),(3,"Saquon Barkley","PHI","17.6"),
      (3,"Ashton Jeanty","LV","18.9"),(3,"Kenneth Walker III","KC","20.5"),(4,"Omarion Hampton","LAC","24.0"),
      (4,"Josh Jacobs","GB","25.4"),(4,"Kyren Williams","LAR","27.4"),(4,"Jeremiyah Love","ARI","29.0"),
      (5,"Javonte Williams","DAL","33.2"),(5,"Breece Hall","NYJ","33.5"),(5,"Travis Etienne Jr.","NO","37.3"),
      (5,"Cam Skattebo","NYG","38.4"),(6,"D'Andre Swift","CHI","43.2"),(6,"Bucky Irving","TB","47.0"),
      (6,"Quinshon Judkins","CLE","51.5"),(6,"Bhayshul Tuten","JAX","52.5"),(6,"David Montgomery","HOU","55.5"),
      (7,"Ollie Gordon II","MIA","-"),(7,"Jaylen Warren","PIT","-"),(7,"R. Stevenson","NE","-"),
      (7,"TreVeyon Henderson","NE","-"),(8,"Tony Pollard","TEN","-"),(8,"Rico Dowdle","PIT","-"),
      (8,"J.K. Dobbins","DEN","-"),(9,"Chuba Hubbard","CAR","-"),(9,"Jonathon Brooks","CAR","-"),
      (10,"RJ Harvey","DEN","-"),(10,"Kenny Gainwell","TB","-")]

WR = [(1,"Puka Nacua","LAR","3.0"),(1,"Ja'Marr Chase","CIN","3.9"),(1,"J. Smith-Njigba","SEA","5.2"),
      (2,"Amon-Ra St. Brown","DET","8.0"),(2,"Drake London","ATL","12.0"),(2,"CeeDee Lamb","DAL","12.5"),
      (3,"Justin Jefferson","MIN","14.4"),(3,"Rashee Rice","KC","15.2"),(3,"Nico Collins","HOU","18.4"),
      (3,"George Pickens","DAL","19.9"),(3,"A.J. Brown","NE","20.0"),(4,"Chris Olave","NO","23.2"),
      (4,"Zay Flowers","BAL","25.2"),(4,"Malik Nabers","NYG","26.2"),(4,"Tetairoa McMillan","CAR","30.7"),
      (4,"Garrett Wilson","NYJ","31.1"),(4,"DeVonta Smith","PHI","31.6"),(5,"Emeka Egbuka","TB","33.7"),
      (5,"Tee Higgins","CIN","35.9"),(5,"Davante Adams","LAR","37.9"),(5,"Jameson Williams","DET","38.7"),
      (5,"Ladd McConkey","LAC","41.0"),(6,"DJ Moore","BUF","44.9"),(6,"Terry McLaurin","WAS","45.1"),
      (6,"Jaylen Waddle","DEN","47.3"),(6,"Rome Odunze","CHI","48.9"),(6,"Mike Evans","SF","54.2"),
      (7,"Christian Watson","GB","56.0"),(7,"Luther Burden III","CHI","57.6"),(7,"Courtland Sutton","DEN","59.1"),
      (7,"DK Metcalf","PIT","-"),(7,"Marvin Harrison Jr.","ARI","-"),(7,"Brian Thomas Jr.","JAX","-"),
      (8,"M. Pittman Jr.","PIT","-"),(8,"Jordan Addison","MIN","-"),(8,"Jayden Reed","GB","-"),
      (8,"Xavier Worthy","KC","-"),(9,"Matthew Golden","GB","-"),(9,"Khalil Shakir","BUF","-"),
      (9,"Makai Lemon","PHI","-")]

QB = [(5,"Josh Allen","BUF","33.0"),(6,"Drake Maye","NE","51.3"),(6,"Lamar Jackson","BAL","55.5"),
      (7,"Joe Burrow","CIN","57.5"),(8,"Dak Prescott","DAL","-"),(8,"Jayden Daniels","WAS","-"),
      (8,"Matthew Stafford","LAR","-"),(9,"Jalen Hurts","PHI","-"),(9,"Brock Purdy","SF","-"),
      (9,"Trevor Lawrence","JAX","-"),(9,"Caleb Williams","CHI","-"),(9,"Jared Goff","DET","-"),
      (9,"Patrick Mahomes","KC","-"),(10,"Justin Herbert","LAC","-"),(10,"Bo Nix","DEN","-"),
      (10,"Jaxson Dart","NYG","-")]

TE = [(5,"Trey McBride","ARI","37.8"),(6,"Brock Bowers","LV","44.0"),(8,"George Kittle","SF","-"),
      (9,"Colston Loveland","CHI","-"),(9,"Tyler Warren","IND","-"),(9,"Harold Fannin Jr.","CLE","-"),
      (10,"Kyle Pitts Sr.","ATL","-"),(10,"Dallas Goedert","PHI","-")]

wb3.cell(1,1)  # anchor
pos_block(wb3, 1, "RUNNING BACKS", POS_COLOR["RB"], RB)
pos_block(wb3, 6, "WIDE RECEIVERS", POS_COLOR["WR"], WR)
pos_block(wb3, 11, "QUARTERBACKS", POS_COLOR["QB"], QB)
pos_block(wb3, 16, "TIGHT ENDS", POS_COLOR["TE"], TE)
wb3.freeze_panes = "A3"
wb3["U1"] = "ADP '-' = expert-rank tier (beyond live top-60 ADP)"
wb3["U1"].font = F(9, italic=True, color="777777")
qbn = wb3.cell(20, 11, "6-PT PASS TDs — WHAT THE DATA ACTUALLY SAYS: the rule lifts every QB, "
    "streamer included, so the elite QB's edge over a streamer grows only 7-19 pts across a WHOLE season "
    "(median 12.5, measured over 2024+2025 x 3 replacement ranks). The rule pays exactly 2 pts per "
    "PASSING TD and nothing extra for rushing, so judge on passing-TD volume — 2025: Stafford +92, "
    "Goff +68, Maye +62; 2024: Burrow +86, Mayfield +82, Jackson +82. Wait on QB until round 9. "
    "Source: build_edge.py on nflverse.")
qbn.font = F(9, italic=True, color="B45F06"); qbn.alignment = LEFTTOP
wb3.merge_cells(start_row=20, start_column=11, end_row=25, end_column=14)

# ============================================================
# TAB 4 — SLEEPERS, VALUES & FADES
# ============================================================
wb4 = wb.create_sheet("Sleepers • Values • Fades")
wb4.sheet_properties.tabColor = "B45F06"

def section(ws, row, title, color, items):
    tcell = ws.cell(row, 1, title); tcell.font = F(12, bold=True, color=WHITE)
    tcell.fill = fill(color); tcell.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    for (name, note) in items:
        nc = ws.cell(row, 1, name); nc.font = F(10, bold=True); nc.alignment = LEFT; nc.border = BORDER
        note_c = ws.cell(row, 2, note); note_c.font = F(9); note_c.alignment = LEFT; note_c.border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 28
        row += 1
    return row + 1

wb4["A1"] = "SLEEPERS • VALUES • FADES • INJURY WATCH"
wb4["A1"].font = F(15, bold=True, color=NAVY)
wb4.merge_cells("A1:C1")
wb4["A1"].alignment = LEFT
row = 3

row = section(wb4, row, "VALUES (talent > draft cost — target these)", "137333", [
    ("DeVonta Smith (WR, PHI)", "Locked-in WR2 going as a WR3."),
    ("Breece Hall (RB, NYJ)", "Groin now, but RB1 talent at an RB2 price — buy the dip if healthy."),
    ("Jameson Williams (WR, DET)", "Big-play ceiling in a loaded offense."),
    ("TreVeyon Henderson (RB, NE)", "Explosive; can seize the NE backfield."),
    ("Matthew Stafford / Dak Prescott (QB)", "46 and 30 pass TDs in 2025 = +92 and +60 pts from the 6-pt rule, going 50+ picks after Allen. Judge QBs on passing-TD volume, not rushing reputation."),
    ("Marvin Harrison Jr. (WR, ARI)", "Year-3 bounce-back at a discount."),
])

row = section(wb4, row, "SLEEPERS (late-round upside swings)", "1155CC", [
    ("Makai Lemon (WR, PHI)", "Rookie with real target path."),
    ("Carnell Tate (WR, TEN)", "Rookie; volume if he wins the role."),
    ("Matthew Golden (WR, GB)", "Ascending role in GB."),
    ("Jaylen Warren (RB, PIT)", "Standalone value + league-winner if Harris-type role opens."),
    ("Jalen Coker (WR, CAR)", "Cheap target share."),
    ("Jaxson Dart (QB, NYG)", "Late-round rushing-QB lottery ticket."),
])

row = section(wb4, row, "INJURY WATCH (draftable — confirm status draft day)", "B45F06", [
    ("Ashton Jeanty (RB, LV)", "Ankle — Wk1 uncertain; ADP slid ~15 to 19. Real discount if it drops him to your R2 turn."),
    ("Puka Nacua (WR, LAR)", "Groin — expected Week 1."),
    ("Malik Nabers (WR, NYG)", "Knee — expected Week 1."),
    ("Breece Hall (RB, NYJ)", "Groin, out 2-3 weeks — expected Week 1."),
    ("Emeka Egbuka (WR, TB)", "Toe — day-to-day, monitor."),
    ("Jeremiyah Love (RB, ARI)", "High-ankle sprain — expected Week 1."),
    ("George Kittle (TE, SF)", "Achilles — activated off PUP, expected Week 1."),
    ("Patrick Mahomes (QB, KC)", "Knee — playing, but rushing upside may dip."),
])

row = section(wb4, row, "FADES / AVOID (hurt too much or priced too high)", "CC0000", [
    ("Ricky Pearsall (WR, SF)", "OUT for 2026 (PCL surgery). Do not draft."),
    ("Jayden Higgins (WR, HOU)", "OUT for 2026 (ACL). Do not draft."),
    ("Jordyn Tyson (WR, NO)", "Hamstring — ~2 months / likely IR. Stash only, don't reach."),
    ("Kyle Monangai (RB, CHI)", "Knee — multi-week; falling value."),
    ("Cam Skattebo (RB, NYG)", "Flagged as one of the riskiest top-50 picks."),
])

row = section(wb4, row, "HANDCUFFS (grab late to protect your studs)", "666666", [
    ("RJ Harvey (DEN)", "Behind J.K. Dobbins."),
    ("Jonathon Brooks (CAR)", "Committee with Chuba Hubbard."),
    ("Rico Dowdle (PIT)", "Behind Jaylen Warren."),
    ("Kenny Gainwell (TB)", "Behind Bucky Irving."),
    ("Rhamondre Stevenson / T. Henderson (NE)", "Whichever you don't roster."),
])

wb4.column_dimensions["A"].width = 34
wb4.column_dimensions["B"].width = 30
wb4.column_dimensions["C"].width = 30
wb4.freeze_panes = "A3"

# ============================================================
# TAB 5 — FAAB ($100)
# ============================================================
wb5 = wb.create_sheet("FAAB ($100)")
wb5.sheet_properties.tabColor = "137333"
wb5["A1"] = "FAAB — $100 SEASON FREE-AGENT BUDGET (in-season waiver bidding)"
wb5["A1"].font = F(14, bold=True, color=NAVY)
wb5.merge_cells("A1:B1")
wb5["A2"] = "This is your in-season pot for bidding on free agents — separate from the draft. How to make it last:"
wb5["A2"].font = F(9, italic=True, color="555555")
wb5.merge_cells("A2:B2")
faab = [
    ("Save it for RBs", "Most FAAB league-winners are backup RBs who fall into a starting job (injury or benching). Spend your biggest bids there, not on WRs."),
    ("Be aggressive early (Wks 1-4)", "This is when breakout RB/WRs first surface and roles shift. Pounce on real opportunity — but don't empty the tank."),
    ("Cap any single bid", "Rarely go over ~$30-40 on one player unless it's a clear league-winner (a true every-down bell-cow RB)."),
    ("Keep dry powder", "Hold ~$25-30 into the fantasy playoff push (Wks 12-14) for an injury replacement or a great matchup streamer."),
    ("Win ties with odd bids", "Bid $17 or $23, never round numbers — you edge out everyone parked on $15 / $20."),
    ("Stream defenses cheap", "No kicker in this league. Defenses and one-week fill-ins are $0-2 bids — stream the matchup, never pay up."),
    ("Don't panic Week 1", "Let roles settle a week or two before spending on preseason hype or one big game."),
]
r = 4
for (h, note) in faab:
    hc = wb5.cell(r, 1, h); hc.font = F(10, bold=True); hc.alignment = LEFTTOP; hc.fill = fill("EAF3EA"); hc.border = BORDER
    nc = wb5.cell(r, 2, note); nc.font = F(10); nc.alignment = LEFTTOP; nc.border = BORDER
    wb5.row_dimensions[r].height = 32
    r += 1
wb5.column_dimensions["A"].width = 26
wb5.column_dimensions["B"].width = 84
wb5.freeze_panes = "A4"

wb.save("fantasy_draft_2026.xlsx")
print("saved fantasy_draft_2026.xlsx")
print("sheets:", wb.sheetnames)
print("big board players:", len(board))

# ---- also emit a CSV for the native Google Sheet (with live tracker formulas) ----
import csv

rows = []
TRACKER_MARK = "__TRACKER__"

rows.append(["2026 DRAFT BOARD — GovSmart Gridiron — Pick #3 (LIVE TRACKER)"])
rows.append(["VERIFIED from Sleeper: 12-team snake | 0.5 PPR | 6-pt passing TDs | $100 FAAB | slot #3 | draft Fri Sep 4 5:30 PM ET | 90s pick timer"])
rows.append(["ROSTER: QB/RB/RB/WR/WR/TE/FLEX/DEF + 4 bench. 12 rounds. NO KICKER. A DEFENSE is a required starter. 2 IR slots."])
rows.append(["6-pt pass TDs lift EVERY QB — take a top-6 QB by rounds 5-8."])
rows.append(["Your 12 picks (snake, slot 3): 3, 22, 27, 46, 51, 70, 75, 94, 99, 118, 123, 142"])
rows.append([])
rows.append(["LIVE TRACKER — mark x in Drafted? (col I) for EVERY pick; ALSO x in MINE? (col J) for yours. Counters update live."])
HDR_POS_ROW = len(rows) + 1                       # 1-indexed row of the QB/RB/WR/TE header
rows.append(["", "QB", "RB", "WR", "TE"])
ROW_COUNT = len(rows) + 1
rows.append([TRACKER_MARK, "count"])
ROW_NEED = len(rows) + 1
rows.append([TRACKER_MARK, "need"])
ROW_ELITE = len(rows) + 1
rows.append([TRACKER_MARK, "elite"])
ROW_MID = len(rows) + 1
rows.append([TRACKER_MARK, "mid"])
BYE_ROW = len(rows) + 1
rows.append(["My bye weeks:", 5, 6, 7, 8, 9, 10, 11, 13, 14])
rows.append([TRACKER_MARK, "byecount"])
rows.append([])
rows.append(["ROUND-BY-ROUND — BEST PLAYER TO TARGET AT EACH PICK"])
rows.append(["Rd", "Your Pick", "On Board (ADP)", "BEST TARGET (take one)", "Fallbacks / Best Available", "Plan"])
for row in plan:
    rows.append(list(row))
rows.append([])
rows.append(["TURN SCENARIOS — PICKS 22 & 27 (you pick back-to-back)"])
for sc in scenarios:
    rows.append([sc])
rows.append([])
rows.append(["BIG BOARD — Top 120 (half-PPR). ADP thru ~#60 = live market (Aug 27); 61+ = expert consensus."])
rows.append(["Rank","Tier","Player","Pos","Team","Bye","ADP","Notes","Drafted?","MINE?"])
B0 = len(rows) + 1                                 # 1-indexed first board data row
for i, (player, pos, team, bye, adp, tier, notes) in enumerate(board, 1):
    rows.append([i, "T%d" % tier, player, pos, team, bye, adp, notes, "", ""])
B1 = len(rows)                                     # 1-indexed last board data row
rows.append([])
rows.append(["FAAB — $100 SEASON FREE-AGENT BUDGET (in-season waiver bidding; separate from the draft)"])
for (h, note) in faab:
    rows.append([h, note])

# --- now that B0/B1 are known, generate the tracker formulas against real rows ---
POSR  = "$D$%d:$D$%d" % (B0, B1); TIERR = "$B$%d:$B$%d" % (B0, B1)
BYER  = "$F$%d:$F$%d" % (B0, B1); DRR   = "$I$%d:$I$%d" % (B0, B1)
MINER = "$J$%d:$J$%d" % (B0, B1)
H = HDR_POS_ROW
STARTERS = {"QB": 1, "RB": 2, "WR": 2, "TE": 1}    # DEF handled separately (R12)

def mine(c):  return '=COUNTIFS(%s,%s$%d,%s,"x")' % (POSR, c, H, MINER)
def band(c, tiers):
    return "=" + "+".join('COUNTIFS(%s,%s$%d,%s,"T%d",%s,"<>x")' % (POSR, c, H, TIERR, t, DRR)
                          for t in tiers)
# position-relative tier bands, same logic as the console
bands = {}
for pos in ("QB","RB","WR","TE"):
    ts = sorted({b[5] for b in board if b[1] == pos})
    bands[pos] = (ts[:2], ts[2:4])

COLS = {"QB":"B","RB":"C","WR":"D","TE":"E"}
tracker = {
  "count":   ["My roster count"]        + [mine(COLS[p]) for p in ("QB","RB","WR","TE")],
  "need":    ["Starters still needed"]  + ["=MAX(0,%d-%s%d)" % (STARTERS[p], COLS[p], ROW_COUNT)
                                            for p in ("QB","RB","WR","TE")] + ["(+1 FLEX, +1 DEF in R12)"],
  "elite":   ["Top-tier left (undrafted)"] + [band(COLS[p], bands[p][0]) for p in ("QB","RB","WR","TE")],
  "mid":     ["Next tier left"]         + [band(COLS[p], bands[p][1]) for p in ("QB","RB","WR","TE")],
  "byecount":["Count (3+ same bye = fix it)"] + ['=COUNTIFS(%s,%s$%d,%s,"x")' % (BYER, c, BYE_ROW, MINER)
                                                 for c in "BCDEFGHIJ"],
}
for i, r in enumerate(rows):
    if r and r[0] == TRACKER_MARK:
        rows[i] = tracker[r[1]]

assert rows[B0-2][0] == "Rank", "board header misplaced: %r" % (rows[B0-2][:3],)
assert rows[B0-1][2] == board[0][0], "board start wrong: %r" % (rows[B0-1][:3],)
assert rows[B1-1][2] == board[-1][0], "board end wrong: %r" % (rows[B1-1][:3],)
assert B1 - B0 + 1 == len(board), "board row span %d != %d players" % (B1-B0+1, len(board))
assert not any(r and r[0] == TRACKER_MARK for r in rows), "unreplaced tracker row"

# The offsets above being right is not the same as the FORMULAS being right.
# Parse every generated formula back out and prove each range and each criteria
# cell lands on the row we intended. This is the assertion that was missing when
# a previous version silently pointed every formula at the wrong cells.
import re as _re
_HDRCOLS = ["Rank","Tier","Player","Pos","Team","Bye","ADP","Notes","Drafted?","MINE?"]
assert rows[B0-2] == _HDRCOLS, "board columns moved: %r" % (rows[B0-2],)
_checked = 0
for _r in rows:
    for _c in _r:
        if not (isinstance(_c, str) and _c.startswith("=COUNTIFS")):
            continue
        for _col, _a, _b in _re.findall(r"\$([A-J])\$(\d+):\$[A-J]\$(\d+)", _c):
            assert (int(_a), int(_b)) == (B0, B1), \
                "formula range $%s$%s:$%s$%s != board rows %d-%d in %r" % (
                    _col, _a, _col, _b, B0, B1, _c)
            _checked += 1
        # criteria cells are the bare (unanchored-column) refs like B$8 / B$13
        for _cell, _row in _re.findall(r"(?<![$:\w])([A-J])\$(\d+)", _c):
            assert int(_row) in (HDR_POS_ROW, BYE_ROW), \
                "criteria cell %s$%s points at row %s, expected header %d or %d in %r" % (
                    _cell, _row, _row, HDR_POS_ROW, BYE_ROW, _c)
            _checked += 1
assert _checked >= 30, "only %d formula refs verified — assertions not running" % _checked
assert rows[HDR_POS_ROW-1][1:5] == ["QB","RB","WR","TE"], \
    "position header row %d moved: %r" % (HDR_POS_ROW, rows[HDR_POS_ROW-1][:5])
assert len([c for c in rows[BYE_ROW-1][1:] if c != ""]) == 9, \
    "bye header row %d should hold 9 weeks: %r" % (BYE_ROW, rows[BYE_ROW-1])
assert len([c for c in rows[BYE_ROW][1:] if str(c).startswith("=")]) == 9, \
    "bye count row should hold 9 formulas: %r" % (rows[BYE_ROW],)
print("verified %d formula cell references against computed offsets" % _checked)
print("CSV: board rows %d-%d | tracker rows %d-%d | bands %s" % (B0, B1, ROW_COUNT, BYE_ROW+1,
      {p: bands[p][0] for p in bands}))

with open("draft_sheet.csv", "w", newline="") as f:
    csv.writer(f).writerows(rows)
print("wrote draft_sheet.csv")
